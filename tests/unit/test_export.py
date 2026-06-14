"""
Unit tests for the export module.

Tests JSON and Excel exporters with comprehensive coverage
of configuration options and edge cases.
"""

import json
import tempfile
from pathlib import Path
from typing import Any

import pytest

from src.export.excel_exporter import (
    ExcelExportConfig,
    ExcelExporter,
    ExcelStyler,
    SheetConfig,
    SheetType,
    export_to_excel,
)
from src.export.json_exporter import (
    ExportFormat,
    JSONExportConfig,
    JSONExporter,
    export_to_json,
)
from src.pipeline.state import ConfidenceLevel, ExtractionStatus


@pytest.fixture
def sample_extraction_state() -> dict[str, Any]:
    """Create a sample extraction state for testing."""
    return {
        "processing_id": "test-proc-001",
        "pdf_path": "/test/sample.pdf",
        "pdf_hash": "abc123def456",
        "document_type": "CMS-1500",
        "selected_schema_name": "cms1500_v1",
        "status": ExtractionStatus.COMPLETED.value,
        "start_time": "2024-01-15T10:00:00Z",
        "end_time": "2024-01-15T10:00:30Z",
        "total_vlm_calls": 4,
        "total_processing_time_ms": 30000,
        "retry_count": 0,
        "overall_confidence": 0.92,
        "confidence_level": ConfidenceLevel.HIGH.value,
        "requires_human_review": False,
        "human_review_reason": "",
        "page_images": [b"page1", b"page2"],
        "merged_extraction": {
            "patient_name": {
                "value": "John Doe",
                "confidence": 0.95,
                "location": "Box 2",
            },
            "date_of_service": {
                "value": "01/15/2024",
                "confidence": 0.88,
                "location": "Box 24A",
            },
            "total_charges": {
                "value": "150.00",
                "confidence": 0.92,
                "location": "Box 28",
            },
            "member_id": {
                "value": "MEM123456",
                "confidence": 0.90,
                "location": "Box 1a",
            },
        },
        "field_metadata": {
            "patient_name": {
                "confidence": 0.95,
                "confidence_level": "high",
                "passes_agree": True,
                "validation_passed": True,
            },
            "date_of_service": {
                "confidence": 0.88,
                "confidence_level": "high",
                "passes_agree": True,
                "validation_passed": True,
            },
            "total_charges": {
                "confidence": 0.92,
                "confidence_level": "high",
                "passes_agree": True,
                "validation_passed": True,
            },
            "member_id": {
                "confidence": 0.90,
                "confidence_level": "high",
                "passes_agree": True,
                "validation_passed": True,
            },
        },
        "validation": {
            "is_valid": True,
            "field_validations": {
                "patient_name": {"is_valid": True, "validation_type": "format"},
                "date_of_service": {"is_valid": True, "validation_type": "date"},
            },
            "cross_field_validations": [],
            "hallucination_flags": [],
            "warnings": [],
            "errors": [],
        },
        "page_extractions": [
            {
                "page_number": 1,
                "merged_fields": {"patient_name": "John Doe"},
                "overall_confidence": 0.93,
                "agreement_rate": 0.95,
                "vlm_calls": 2,
                "extraction_time_ms": 15000,
                "errors": [],
                "pass1_raw": {"patient_name": {"value": "John Doe", "confidence": 0.94}},
                "pass2_raw": {"patient_name": {"value": "John Doe", "confidence": 0.95}},
            },
            {
                "page_number": 2,
                "merged_fields": {"total_charges": "150.00"},
                "overall_confidence": 0.91,
                "agreement_rate": 0.92,
                "vlm_calls": 2,
                "extraction_time_ms": 15000,
                "errors": [],
                "pass1_raw": {"total_charges": {"value": "150.00", "confidence": 0.91}},
                "pass2_raw": {"total_charges": {"value": "150.00", "confidence": 0.92}},
            },
        ],
        "errors": [],
        "warnings": [],
    }


class TestJSONExporter:
    """Test cases for JSONExporter."""

    def test_default_config(self) -> None:
        """Test default configuration values."""
        config = JSONExportConfig()
        assert config.format == ExportFormat.STANDARD
        assert config.include_metadata is True
        assert config.include_confidence is True
        assert config.pretty_print is True
        assert config.mask_phi is False

    def test_minimal_export(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test minimal export format."""
        config = JSONExportConfig(format=ExportFormat.MINIMAL)
        exporter = JSONExporter(config)

        result = exporter.export(sample_extraction_state)

        assert "data" in result
        assert "processing_id" in result
        assert "status" in result
        assert result["data"]["patient_name"] == "John Doe"
        assert "metadata" not in result
        assert "confidence" not in result

    def test_standard_export(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test standard export format."""
        config = JSONExportConfig(format=ExportFormat.STANDARD)
        exporter = JSONExporter(config)

        result = exporter.export(sample_extraction_state)

        assert "data" in result
        assert "processing_id" in result
        assert "document_type" in result
        assert "confidence" in result
        assert "metadata" in result
        assert result["confidence"]["overall"] == 0.92

    def test_detailed_export(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test detailed export format."""
        config = JSONExportConfig(format=ExportFormat.DETAILED)
        exporter = JSONExporter(config)

        result = exporter.export(sample_extraction_state)

        assert "data" in result
        assert "validation" in result
        assert "pages" in result
        assert "audit" in result
        assert len(result["pages"]) == 2

    def test_fhir_export(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test FHIR-compatible export format."""
        config = JSONExportConfig(format=ExportFormat.FHIR_COMPATIBLE)
        exporter = JSONExporter(config)

        result = exporter.export(sample_extraction_state)

        assert result["resourceType"] == "DocumentReference"
        assert result["id"] == "test-proc-001"
        assert "content" in result
        assert "extension" in result

    def test_phi_masking(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test PHI field masking."""
        config = JSONExportConfig(
            format=ExportFormat.STANDARD,
            mask_phi=True,
            phi_fields={"member_id"},
        )
        exporter = JSONExporter(config)

        result = exporter.export(sample_extraction_state)

        # member_id should be masked
        assert "***MASKED***" in result["data"]["member_id"]
        # patient_name should not be masked
        assert result["data"]["patient_name"] == "John Doe"

    def test_exclude_fields(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test field exclusion."""
        config = JSONExportConfig(
            format=ExportFormat.STANDARD,
            exclude_fields={"patient_name", "date_of_service"},
        )
        exporter = JSONExporter(config)

        result = exporter.export(sample_extraction_state)

        assert "patient_name" not in result["data"]
        assert "date_of_service" not in result["data"]
        assert "total_charges" in result["data"]

    def test_write_to_file(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test writing export to file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "output.json"

            exporter = JSONExporter()
            exporter.export(sample_extraction_state, output_path=output_path)

            assert output_path.exists()

            with open(output_path) as f:
                loaded = json.load(f)

            assert loaded["processing_id"] == "test-proc-001"

    def test_export_to_json_convenience(
        self,
        sample_extraction_state: dict[str, Any],
    ) -> None:
        """Test convenience function."""
        result = export_to_json(
            sample_extraction_state,
            format=ExportFormat.DETAILED,
            include_metadata=True,
        )

        assert "data" in result
        assert "metadata" in result
        assert "validation" in result

    def test_include_raw_passes(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test including raw pass data."""
        config = JSONExportConfig(
            format=ExportFormat.DETAILED,
            include_raw_passes=True,
        )
        exporter = JSONExporter(config)

        result = exporter.export(sample_extraction_state)

        assert "raw_passes" in result
        assert len(result["raw_passes"]) == 2

    def test_empty_state(self) -> None:
        """Test export with minimal state."""
        empty_state: dict[str, Any] = {
            "processing_id": "empty-001",
            "status": ExtractionStatus.PENDING.value,
        }

        exporter = JSONExporter()
        result = exporter.export(empty_state)

        assert result["processing_id"] == "empty-001"
        assert result["data"] == {}


class TestExcelExporter:
    """Test cases for ExcelExporter."""

    def test_default_config(self) -> None:
        """Test default configuration values."""
        config = ExcelExportConfig()
        assert len(config.sheets) == 5
        assert config.include_styling is True
        assert config.include_confidence_colors is True
        assert config.mask_phi is False

    def test_basic_export(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test basic Excel export."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "output.xlsx"

            exporter = ExcelExporter()
            result_path = exporter.export(sample_extraction_state, output_path)

            assert result_path.exists()
            assert result_path.suffix == ".xlsx"

    def test_all_sheets_created(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test that all configured sheets are created."""
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "output.xlsx"

            exporter = ExcelExporter()
            exporter.export(sample_extraction_state, output_path)

            workbook = load_workbook(output_path)
            sheet_names = workbook.sheetnames

            assert "Extracted Data" in sheet_names
            assert "Processing Metadata" in sheet_names
            assert "Validation Results" in sheet_names
            assert "Audit Trail" in sheet_names

    def test_custom_sheets(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test custom sheet configuration."""
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            sheets=[
                SheetConfig(SheetType.DATA, "Data"),
                SheetConfig(SheetType.PAGE_DETAILS, "Pages"),
            ],
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "output.xlsx"

            exporter = ExcelExporter(config)
            exporter.export(sample_extraction_state, output_path)

            workbook = load_workbook(output_path)
            sheet_names = workbook.sheetnames

            assert "Data" in sheet_names
            assert "Pages" in sheet_names
            assert len(sheet_names) == 2

    def test_phi_masking(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test PHI masking in Excel export."""
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            mask_phi=True,
            phi_fields={"member_id"},
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "output.xlsx"

            exporter = ExcelExporter(config)
            exporter.export(sample_extraction_state, output_path)

            workbook = load_workbook(output_path)
            data_sheet = workbook["Extracted Data"]

            # Find member_id row and check value is masked
            masked_found = False
            for row in data_sheet.iter_rows(min_row=2):
                if row[0].value == "member_id":
                    if "***MASKED***" in str(row[1].value):
                        masked_found = True
                    break

            assert masked_found

    def test_no_styling(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test export without styling."""
        config = ExcelExportConfig(
            include_styling=False,
            include_confidence_colors=False,
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "output.xlsx"

            exporter = ExcelExporter(config)
            result_path = exporter.export(sample_extraction_state, output_path)

            assert result_path.exists()

    def test_export_to_excel_convenience(
        self,
        sample_extraction_state: dict[str, Any],
    ) -> None:
        """Test convenience function."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "output.xlsx"

            result_path = export_to_excel(
                sample_extraction_state,
                output_path,
                include_styling=True,
                mask_phi=False,
            )

            assert result_path.exists()

    def test_data_sheet_content(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test data sheet contains correct content."""
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "output.xlsx"

            exporter = ExcelExporter()
            exporter.export(sample_extraction_state, output_path)

            workbook = load_workbook(output_path)
            data_sheet = workbook["Extracted Data"]

            # Check headers
            assert data_sheet.cell(1, 1).value == "Field Name"
            assert data_sheet.cell(1, 2).value == "Value"
            assert data_sheet.cell(1, 3).value == "Confidence"

            # Check some data rows exist
            assert data_sheet.max_row > 1

    def test_metadata_sheet_content(
        self,
        sample_extraction_state: dict[str, Any],
    ) -> None:
        """Test metadata sheet contains correct content."""
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "output.xlsx"

            exporter = ExcelExporter()
            exporter.export(sample_extraction_state, output_path)

            workbook = load_workbook(output_path)
            meta_sheet = workbook["Processing Metadata"]

            # Check Processing ID is present
            proc_id_found = False
            for row in meta_sheet.iter_rows(min_row=2):
                if row[0].value == "Processing ID":
                    assert row[1].value == "test-proc-001"
                    proc_id_found = True
                    break

            assert proc_id_found

    def test_include_page_details(
        self,
        sample_extraction_state: dict[str, Any],
    ) -> None:
        """Test page details sheet."""
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            sheets=[
                SheetConfig(SheetType.PAGE_DETAILS, "Page Details"),
            ],
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "output.xlsx"

            exporter = ExcelExporter(config)
            exporter.export(sample_extraction_state, output_path)

            workbook = load_workbook(output_path)
            page_sheet = workbook["Page Details"]

            # Should have 2 page rows plus header
            assert page_sheet.max_row == 3

    def test_include_raw_passes(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test raw passes sheet."""
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            sheets=[
                SheetConfig(SheetType.RAW_PASSES, "Raw Passes"),
            ],
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "output.xlsx"

            exporter = ExcelExporter(config)
            exporter.export(sample_extraction_state, output_path)

            workbook = load_workbook(output_path)
            passes_sheet = workbook["Raw Passes"]

            # Should have header row plus data rows
            assert passes_sheet.max_row > 1


class TestExcelStyler:
    """Test cases for ExcelStyler."""

    def test_header_style_application(self) -> None:
        """Test header styling."""
        from openpyxl import Workbook

        styler = ExcelStyler()
        workbook = Workbook()
        worksheet = workbook.active

        cell = worksheet.cell(1, 1, value="Test Header")
        styler.apply_header_style(cell)

        assert cell.font.bold is True
        assert cell.fill.fill_type == "solid"

    def test_confidence_color_high(self) -> None:
        """Test high confidence coloring."""
        from openpyxl import Workbook

        styler = ExcelStyler()
        workbook = Workbook()
        worksheet = workbook.active

        cell = worksheet.cell(1, 1, value=0.95)
        styler.apply_confidence_color(cell, 0.95)

        assert cell.fill.fill_type == "solid"
        # HIGH_CONFIDENCE_BG is "C6EFCE", but openpyxl may add alpha channel
        assert ExcelStyler.HIGH_CONFIDENCE_BG in cell.fill.start_color.rgb

    def test_confidence_color_medium(self) -> None:
        """Test medium confidence coloring."""
        from openpyxl import Workbook

        styler = ExcelStyler()
        workbook = Workbook()
        worksheet = workbook.active

        cell = worksheet.cell(1, 1, value=0.70)
        styler.apply_confidence_color(cell, 0.70)

        assert cell.fill.fill_type == "solid"
        assert ExcelStyler.MEDIUM_CONFIDENCE_BG in cell.fill.start_color.rgb

    def test_confidence_color_low(self) -> None:
        """Test low confidence coloring."""
        from openpyxl import Workbook

        styler = ExcelStyler()
        workbook = Workbook()
        worksheet = workbook.active

        cell = worksheet.cell(1, 1, value=0.30)
        styler.apply_confidence_color(cell, 0.30)

        assert cell.fill.fill_type == "solid"
        assert ExcelStyler.LOW_CONFIDENCE_BG in cell.fill.start_color.rgb

    def test_status_coloring(self) -> None:
        """Test status-based coloring."""
        from openpyxl import Workbook

        styler = ExcelStyler()
        workbook = Workbook()
        worksheet = workbook.active

        cell = worksheet.cell(1, 1, value="completed")
        styler.apply_status_color(cell, ExtractionStatus.COMPLETED.value)

        assert cell.fill.fill_type == "solid"


class TestExportFormats:
    """Test export format enumeration."""

    def test_format_values(self) -> None:
        """Test export format enum values."""
        assert ExportFormat.MINIMAL.value == "minimal"
        assert ExportFormat.STANDARD.value == "standard"
        assert ExportFormat.DETAILED.value == "detailed"
        assert ExportFormat.FHIR_COMPATIBLE.value == "fhir_compatible"


class TestSheetTypes:
    """Test sheet type enumeration."""

    def test_sheet_type_values(self) -> None:
        """Test sheet type enum values."""
        assert SheetType.DATA.value == "data"
        assert SheetType.METADATA.value == "metadata"
        assert SheetType.VALIDATION.value == "validation"
        assert SheetType.AUDIT.value == "audit"
        assert SheetType.PAGE_DETAILS.value == "page_details"
        assert SheetType.RAW_PASSES.value == "raw_passes"


class TestMarkdownExporter:
    """Test cases for Markdown exporter."""

    def test_default_config(self) -> None:
        """Test default markdown export configuration."""
        from src.export import MarkdownExportConfig, MarkdownStyle

        config = MarkdownExportConfig()

        assert config.style == MarkdownStyle.DETAILED
        assert config.include_toc is True
        assert config.include_confidence_indicators is True
        assert config.mask_phi is False

    def test_custom_config(self) -> None:
        """Test custom markdown export configuration."""
        from src.export import MarkdownExportConfig, MarkdownStyle

        config = MarkdownExportConfig(
            style=MarkdownStyle.SIMPLE,
            include_toc=False,
            mask_phi=True,
        )

        assert config.style == MarkdownStyle.SIMPLE
        assert config.include_toc is False
        assert config.mask_phi is True

    def test_simple_export(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test simple markdown export."""
        from src.export import MarkdownExportConfig, MarkdownExporter, MarkdownStyle

        config = MarkdownExportConfig(style=MarkdownStyle.SIMPLE)
        exporter = MarkdownExporter(config)
        result = exporter.export(sample_extraction_state)

        assert "CMS-1500" in result
        assert "Extracted Data" in result
        assert "patient_name" in result.lower() or "Patient Name" in result

    def test_detailed_export(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test detailed markdown export."""
        from src.export import MarkdownExportConfig, MarkdownExporter, MarkdownStyle

        config = MarkdownExportConfig(style=MarkdownStyle.DETAILED)
        exporter = MarkdownExporter(config)
        result = exporter.export(sample_extraction_state)

        assert "Extraction Report" in result
        assert "Executive Summary" in result
        assert "Processing Metadata" in result

    def test_summary_export(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test summary markdown export."""
        from src.export import MarkdownExportConfig, MarkdownExporter, MarkdownStyle

        config = MarkdownExportConfig(style=MarkdownStyle.SUMMARY)
        exporter = MarkdownExporter(config)
        result = exporter.export(sample_extraction_state)

        assert "Summary" in result
        assert "Status" in result or "status" in result

    def test_technical_export(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test technical markdown export."""
        from src.export import MarkdownExportConfig, MarkdownExporter, MarkdownStyle

        config = MarkdownExportConfig(style=MarkdownStyle.TECHNICAL)
        exporter = MarkdownExporter(config)
        result = exporter.export(sample_extraction_state)

        assert "Technical" in result or "Performance" in result

    def test_phi_masking(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test PHI masking in markdown export."""
        from src.export import MarkdownExportConfig, MarkdownExporter, MarkdownStyle

        config = MarkdownExportConfig(
            style=MarkdownStyle.SIMPLE,
            mask_phi=True,
            phi_fields={"patient_name"},
        )
        exporter = MarkdownExporter(config)
        result = exporter.export(sample_extraction_state)

        # PHI should be masked
        assert "[REDACTED]" in result or "Jo[REDACTED]hn" in result

    def test_export_to_file(
        self,
        sample_extraction_state: dict[str, Any],
        tmp_path: Path,
    ) -> None:
        """Test markdown export to file."""
        from src.export import MarkdownStyle, export_to_markdown

        output_path = tmp_path / "output.md"
        result = export_to_markdown(
            sample_extraction_state,
            output_path=output_path,
            style=MarkdownStyle.DETAILED,
        )

        assert output_path.exists()
        assert output_path.suffix == ".md"

        content = output_path.read_text(encoding="utf-8")
        assert len(content) > 0
        assert "CMS-1500" in content

    def test_confidence_indicators(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test confidence indicator emojis."""
        from src.export import MarkdownExportConfig, MarkdownExporter, MarkdownStyle

        config = MarkdownExportConfig(
            style=MarkdownStyle.DETAILED,
            include_confidence_indicators=True,
        )
        exporter = MarkdownExporter(config)
        result = exporter.export(sample_extraction_state)

        # Should contain emoji indicators
        assert "" in result or "" in result or "" in result

    def test_table_of_contents(self, sample_extraction_state: dict[str, Any]) -> None:
        """Test table of contents generation."""
        from src.export import MarkdownExportConfig, MarkdownExporter, MarkdownStyle

        config = MarkdownExportConfig(
            style=MarkdownStyle.DETAILED,
            include_toc=True,
        )
        exporter = MarkdownExporter(config)
        result = exporter.export(sample_extraction_state)

        assert "Table of Contents" in result


class TestMarkdownStyle:
    """Test markdown style enumeration."""

    def test_style_values(self) -> None:
        """Test markdown style enum values."""
        from src.export import MarkdownStyle

        assert MarkdownStyle.SIMPLE.value == "simple"
        assert MarkdownStyle.DETAILED.value == "detailed"
        assert MarkdownStyle.SUMMARY.value == "summary"
        assert MarkdownStyle.TECHNICAL.value == "technical"
