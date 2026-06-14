"""Sanity tests for Excel provenance sheet (Phase 8.5-A6).

Verifies that:
  * ``ExcelExportConfig.from_style`` produces the right roster.
  * Default config preserves the legacy 5-sheet shape (no provenance).
  * MINIMAL / SUMMARY require explicit opt-in for provenance.
  * DETAILED / TECHNICAL auto-include the provenance sheet.
  * The provenance sheet has the documented columns and writes one
    row per ``merged_extraction_v2`` field.
  * The data sheet gains a ``_provenance_ref`` column matching the
    provenance sheet's ``field_path`` keys.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import load_workbook

from src.export.excel_exporter import (
    ExcelExportConfig,
    ExcelExporter,
    ExcelStyle,
    SheetType,
)
from src.pipeline.provenance import FieldValue, Provenance


def _build_state() -> dict:
    prov = Provenance(
        page=1,
        source_block_id="blk_p1_001",
        extraction_path=["pass1_vlm", "reconciler"],
        agent_signatures=["extractor", "validator"],
        confidence=0.91,
        vlm_model_id="qwen3.6-27b-vl@8001",
    )
    return {
        "processing_id": "p1",
        "status": "completed",
        "merged_extraction": {"patient_name": {"value": "John", "confidence": 0.9}},
        "merged_extraction_v2": {
            "patient_name": FieldValue(value="John", provenance=prov)
        },
    }


def test_default_config_unchanged() -> None:
    c = ExcelExportConfig()
    assert len(c.sheets) == 5
    assert c.include_provenance is False
    assert SheetType.PROVENANCE not in {s.sheet_type for s in c.sheets}


def test_from_style_minimal_no_provenance() -> None:
    c = ExcelExportConfig.from_style(ExcelStyle.MINIMAL)
    assert len(c.sheets) == 1
    assert c.include_provenance is False


def test_from_style_summary_no_provenance() -> None:
    c = ExcelExportConfig.from_style(ExcelStyle.SUMMARY)
    assert len(c.sheets) == 2
    assert c.include_provenance is False


def test_from_style_detailed_with_provenance() -> None:
    c = ExcelExportConfig.from_style(ExcelStyle.DETAILED)
    assert len(c.sheets) == 6
    assert c.include_provenance is True
    assert c.sheets[-1].sheet_type == SheetType.PROVENANCE


def test_from_style_technical_with_provenance() -> None:
    c = ExcelExportConfig.from_style(ExcelStyle.TECHNICAL)
    assert len(c.sheets) == 8
    assert c.include_provenance is True
    assert c.sheets[-1].sheet_type == SheetType.PROVENANCE


def test_from_style_minimal_explicit_optin() -> None:
    c = ExcelExportConfig.from_style(ExcelStyle.MINIMAL, include_provenance=True)
    assert len(c.sheets) == 2
    assert c.include_provenance is True
    assert c.sheets[-1].sheet_type == SheetType.PROVENANCE


def test_provenance_sheet_has_expected_columns() -> None:
    state = _build_state()
    c = ExcelExportConfig.from_style(ExcelStyle.DETAILED)
    with tempfile.TemporaryDirectory() as d:
        out = Path(d) / "x.xlsx"
        ExcelExporter(c).export(state, out)
        wb = load_workbook(out)
        assert "Provenance" in wb.sheetnames
        p = wb["Provenance"]
        headers = [cell.value for cell in p[1]]
        assert headers == [
            "field_path",
            "page",
            "bbox_x",
            "bbox_y",
            "bbox_width",
            "bbox_height",
            "source_block_id",
            "extraction_path",
            "agent_signatures",
            "confidence",
            "vlm_model_id",
            "mem0_match",
        ]
        # Row 2 corresponds to patient_name
        assert p.cell(2, 1).value == "patient_name"
        assert p.cell(2, 2).value == 1  # page
        assert p.cell(2, 7).value == "blk_p1_001"
        assert "pass1_vlm" in str(p.cell(2, 8).value)
        assert "extractor" in str(p.cell(2, 9).value)
        assert abs(float(p.cell(2, 10).value) - 0.91) < 1e-6
        assert p.cell(2, 11).value == "qwen3.6-27b-vl@8001"


def test_data_sheet_has_provenance_ref_when_enabled() -> None:
    state = _build_state()
    c = ExcelExportConfig.from_style(ExcelStyle.DETAILED)
    with tempfile.TemporaryDirectory() as d:
        out = Path(d) / "x.xlsx"
        ExcelExporter(c).export(state, out)
        wb = load_workbook(out)
        data = wb["Extracted Data"]
        last_header = data.cell(1, data.max_column).value
        assert last_header == "_provenance_ref"
        # Locate the patient_name row.
        for row in data.iter_rows(min_row=2, values_only=True):
            if row[0] == "patient_name":
                assert row[-1] == "patient_name"
                break
        else:
            raise AssertionError("patient_name row missing from data sheet")


def test_data_sheet_no_provenance_ref_when_disabled() -> None:
    """Default config (no provenance) must keep the legacy header set."""
    state = _build_state()
    c = ExcelExportConfig()
    with tempfile.TemporaryDirectory() as d:
        out = Path(d) / "x.xlsx"
        ExcelExporter(c).export(state, out)
        wb = load_workbook(out)
        data = wb["Extracted Data"]
        headers = [cell.value for cell in data[1]]
        assert "_provenance_ref" not in headers
        assert "Provenance" not in wb.sheetnames


def test_provenance_sheet_empty_when_v2_missing() -> None:
    """No merged_extraction_v2 → provenance sheet has header only."""
    state = {
        "processing_id": "p1",
        "status": "completed",
        "merged_extraction": {"patient_name": {"value": "John", "confidence": 0.9}},
    }
    c = ExcelExportConfig.from_style(ExcelStyle.DETAILED)
    with tempfile.TemporaryDirectory() as d:
        out = Path(d) / "x.xlsx"
        ExcelExporter(c).export(state, out)
        wb = load_workbook(out)
        p = wb["Provenance"]
        assert p.max_row == 1  # header only
