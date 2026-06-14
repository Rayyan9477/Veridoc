"""
Unit tests for Phase 4A: Enhanced Export Pipeline Intelligence.

Tests that Excel, JSON, and Markdown exporters correctly surface
Phase 2A-3C pipeline metadata (document splitting, table detection,
schema proposal, prompt enhancement, extraction mode, memory context).
"""

from __future__ import annotations

import json
from typing import Any

import pytest

from src.export.excel_exporter import (
    ExcelExportConfig,
    ExcelExporter,
    SheetConfig,
    SheetType,
    export_to_excel,
)
from src.export.json_exporter import (
    ExportFormat,
    JSONExportConfig,
    JSONExporter,
    export_to_json,
)
from src.export.markdown_exporter import (
    MarkdownExportConfig,
    MarkdownExporter,
    MarkdownStyle,
    export_to_markdown,
)
from src.pipeline.state import (
    ConfidenceLevel,
    ExtractionStatus,
    create_initial_state,
)


# ──────────────────────────────────────────────────────────────────
# Fixtures
# ──────────────────────────────────────────────────────────────────


@pytest.fixture
def base_state() -> dict[str, Any]:
    """Minimal extraction state with Phase 2A-3C fields."""
    return {
        "processing_id": "test-enhanced-001",
        "pdf_path": "/test/enhanced.pdf",
        "pdf_hash": "hash123",
        "document_type": "CMS-1500",
        "selected_schema_name": "cms1500_v1",
        "status": ExtractionStatus.COMPLETED.value,
        "start_time": "2025-01-15T10:00:00Z",
        "end_time": "2025-01-15T10:00:30Z",
        "total_vlm_calls": 6,
        "total_processing_time_ms": 45000,
        "retry_count": 0,
        "overall_confidence": 0.88,
        "confidence_level": ConfidenceLevel.HIGH.value,
        "requires_human_review": False,
        "human_review_reason": "",
        "page_images": [{"page_number": 1}],
        "merged_extraction": {
            "patient_name": {"value": "John Doe", "location": "top-left"},
            "total_charges": {"value": "150.00", "location": "bottom-right"},
        },
        "field_metadata": {
            "patient_name": {
                "confidence": 0.95,
                "confidence_level": "high",
                "passes_agree": True,
                "validation_passed": True,
            },
            "total_charges": {
                "confidence": 0.80,
                "confidence_level": "medium",
                "passes_agree": False,
                "validation_passed": True,
            },
        },
        "validation": {
            "is_valid": True,
            "field_validations": {},
            "cross_field_validations": [],
            "hallucination_flags": [],
        },
        "page_extractions": [],
        "errors": [],
        "warnings": [],
        # Phase 2A fields
        "is_multi_document": False,
        "document_segments": [],
        "active_segment_index": 0,
        # Phase 2B fields
        "detected_tables": [],
        # Phase 2C fields
        "schema_proposal": None,
        # Phase 3B fields
        "prompt_enhancement_applied": False,
        # VLM-first fields
        "use_adaptive_extraction": True,
        "layout_analyses": [],
        "component_maps": [],
        "adaptive_schema": None,
        # Memory fields
        "similar_docs": [],
        "correction_hints": None,
        "provider_patterns": None,
    }


@pytest.fixture
def rich_pipeline_state(base_state: dict[str, Any]) -> dict[str, Any]:
    """State with all pipeline intelligence features populated."""
    base_state.update(
        {
            # Phase 2A: Multi-document with segments
            "is_multi_document": True,
            "document_segments": [
                {"start_page": 1, "end_page": 3, "document_type": "CMS-1500"},
                {"start_page": 4, "end_page": 5, "document_type": "EOB"},
            ],
            # Phase 2B: Tables detected
            "detected_tables": [
                {"page": 1, "row_count": 5, "column_count": 4},
                {"page": 3, "rows": 8, "columns": 6},
            ],
            # Phase 2C: Schema proposal
            "schema_proposal": {
                "schema_name": "auto_cms1500",
                "fields": [
                    {"field_name": "patient_name"},
                    {"field_name": "dob"},
                    {"field_name": "total_charges"},
                ],
            },
            # Phase 3B: Prompt enhancement applied
            "prompt_enhancement_applied": True,
            # VLM-first with analyses
            "use_adaptive_extraction": True,
            "layout_analyses": [{"page": 1}, {"page": 2}],
            "component_maps": [{"page": 1}],
            "adaptive_schema": {"fields": [{"name": "f1"}]},
            # Memory context
            "similar_docs": ["doc-001", "doc-002"],
            "correction_hints": {"patient_name": {"total_corrections": 3}},
            "provider_patterns": {"provider_x": {"pattern": "format-a"}},
        }
    )
    return base_state


# ──────────────────────────────────────────────────────────────────
# Excel: SheetType.PIPELINE
# ──────────────────────────────────────────────────────────────────


class TestExcelPipelineSheet:
    """Tests for the Pipeline Intelligence sheet in Excel exports."""

    def test_pipeline_sheet_type_exists(self):
        assert SheetType.PIPELINE == "pipeline"

    def test_pipeline_sheet_in_default_config(self):
        config = ExcelExportConfig()
        sheet_types = [s.sheet_type for s in config.sheets]
        assert SheetType.PIPELINE in sheet_types

    def test_pipeline_sheet_exported(self, base_state, tmp_path):
        config = ExcelExportConfig(
            sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence")],
            include_styling=False,
        )
        exporter = ExcelExporter(config)
        out = tmp_path / "test.xlsx"
        exporter.export(base_state, out)
        assert out.exists()

    def test_pipeline_sheet_has_headers(self, base_state, tmp_path):
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence")],
            include_styling=False,
        )
        exporter = ExcelExporter(config)
        out = tmp_path / "test.xlsx"
        exporter.export(base_state, out)

        wb = load_workbook(out)
        ws = wb["Pipeline Intelligence"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 4)]
        assert headers == ["Category", "Property", "Value"]

    def test_pipeline_sheet_contains_splitting_info(self, rich_pipeline_state, tmp_path):
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence")],
            include_styling=False,
        )
        exporter = ExcelExporter(config)
        out = tmp_path / "test.xlsx"
        exporter.export(rich_pipeline_state, out)

        wb = load_workbook(out)
        ws = wb["Pipeline Intelligence"]

        # Collect all rows
        rows = []
        for r in range(2, ws.max_row + 1):
            rows.append(
                (ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value)
            )

        categories = [r[0] for r in rows]
        assert "Document Splitting" in categories

        # Check multi-document flag
        multi_doc_row = next(r for r in rows if r[1] == "Is Multi-Document")
        assert multi_doc_row[2] == "Yes"

        # Check segment count
        seg_count_row = next(r for r in rows if r[1] == "Segment Count")
        assert seg_count_row[2] == "2"

    def test_pipeline_sheet_contains_table_info(self, rich_pipeline_state, tmp_path):
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence")],
            include_styling=False,
        )
        exporter = ExcelExporter(config)
        out = tmp_path / "test.xlsx"
        exporter.export(rich_pipeline_state, out)

        wb = load_workbook(out)
        ws = wb["Pipeline Intelligence"]

        rows = []
        for r in range(2, ws.max_row + 1):
            rows.append(
                (ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value)
            )

        table_row = next(r for r in rows if r[1] == "Tables Detected")
        assert table_row[2] == "2"

    def test_pipeline_sheet_contains_enhancement_info(self, rich_pipeline_state, tmp_path):
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence")],
            include_styling=False,
        )
        exporter = ExcelExporter(config)
        out = tmp_path / "test.xlsx"
        exporter.export(rich_pipeline_state, out)

        wb = load_workbook(out)
        ws = wb["Pipeline Intelligence"]

        rows = []
        for r in range(2, ws.max_row + 1):
            rows.append(
                (ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value)
            )

        enhancement_row = next(r for r in rows if r[1] == "Correction-Based Enhancement")
        assert enhancement_row[2] == "Applied"

    def test_pipeline_sheet_contains_memory_info(self, rich_pipeline_state, tmp_path):
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence")],
            include_styling=False,
        )
        exporter = ExcelExporter(config)
        out = tmp_path / "test.xlsx"
        exporter.export(rich_pipeline_state, out)

        wb = load_workbook(out)
        ws = wb["Pipeline Intelligence"]

        rows = []
        for r in range(2, ws.max_row + 1):
            rows.append(
                (ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value)
            )

        similar_row = next(r for r in rows if r[1] == "Similar Documents Found")
        assert similar_row[2] == "2"

        corrections_row = next(r for r in rows if r[1] == "Correction Hints Available")
        assert corrections_row[2] == "Yes"

    def test_pipeline_sheet_minimal_state(self, base_state, tmp_path):
        """Pipeline sheet still works with no pipeline features active."""
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence")],
            include_styling=False,
        )
        exporter = ExcelExporter(config)
        out = tmp_path / "test.xlsx"
        exporter.export(base_state, out)

        wb = load_workbook(out)
        ws = wb["Pipeline Intelligence"]

        rows = []
        for r in range(2, ws.max_row + 1):
            rows.append(
                (ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value)
            )

        # Should still have rows for each category
        multi_row = next(r for r in rows if r[1] == "Is Multi-Document")
        assert multi_row[2] == "No"

        tables_row = next(r for r in rows if r[1] == "Tables Detected")
        assert tables_row[2] == "0"

    def test_pipeline_sheet_can_be_excluded(self, base_state, tmp_path):
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            sheets=[
                SheetConfig(SheetType.DATA, "Data"),
                SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence", include=False),
            ],
            include_styling=False,
        )
        exporter = ExcelExporter(config)
        out = tmp_path / "test.xlsx"
        exporter.export(base_state, out)

        wb = load_workbook(out)
        assert "Pipeline Intelligence" not in wb.sheetnames

    def test_pipeline_sheet_segment_details(self, rich_pipeline_state, tmp_path):
        """Verify individual segment rows appear."""
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence")],
            include_styling=False,
        )
        exporter = ExcelExporter(config)
        out = tmp_path / "test.xlsx"
        exporter.export(rich_pipeline_state, out)

        wb = load_workbook(out)
        ws = wb["Pipeline Intelligence"]

        rows = []
        for r in range(2, ws.max_row + 1):
            rows.append(
                (ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value)
            )

        seg1 = next(r for r in rows if r[1] == "Segment 1")
        assert "Pages 1-3" in seg1[2]
        assert "CMS-1500" in seg1[2]

        seg2 = next(r for r in rows if r[1] == "Segment 2")
        assert "Pages 4-5" in seg2[2]

    def test_pipeline_sheet_schema_proposal(self, rich_pipeline_state, tmp_path):
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence")],
            include_styling=False,
        )
        exporter = ExcelExporter(config)
        out = tmp_path / "test.xlsx"
        exporter.export(rich_pipeline_state, out)

        wb = load_workbook(out)
        ws = wb["Pipeline Intelligence"]

        rows = []
        for r in range(2, ws.max_row + 1):
            rows.append(
                (ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value)
            )

        proposal_row = next(r for r in rows if r[1] == "Proposal Generated")
        assert proposal_row[2] == "Yes"

        name_row = next(r for r in rows if r[1] == "Proposed Schema Name")
        assert name_row[2] == "auto_cms1500"

    def test_pipeline_sheet_adaptive_mode(self, rich_pipeline_state, tmp_path):
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence")],
            include_styling=False,
        )
        exporter = ExcelExporter(config)
        out = tmp_path / "test.xlsx"
        exporter.export(rich_pipeline_state, out)

        wb = load_workbook(out)
        ws = wb["Pipeline Intelligence"]

        rows = []
        for r in range(2, ws.max_row + 1):
            rows.append(
                (ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value)
            )

        adaptive_row = next(r for r in rows if r[1] == "Adaptive (VLM-First)")
        assert adaptive_row[2] == "Yes"

        layout_row = next(r for r in rows if r[1] == "Layout Analyses")
        assert layout_row[2] == "2"

        schema_row = next(r for r in rows if r[1] == "Adaptive Schema Generated")
        assert schema_row[2] == "Yes"


# ──────────────────────────────────────────────────────────────────
# JSON: Pipeline Intelligence Section
# ──────────────────────────────────────────────────────────────────


class TestJSONPipelineIntelligence:
    """Tests for pipeline intelligence in JSON exports."""

    def test_standard_format_has_pipeline_key(self, rich_pipeline_state):
        exporter = JSONExporter(JSONExportConfig(format=ExportFormat.STANDARD))
        result = exporter.export(rich_pipeline_state)
        assert "pipeline" in result

    def test_minimal_format_has_pipeline_key(self, rich_pipeline_state):
        """Minimal format should NOT have pipeline since it only has data + ID + status."""
        exporter = JSONExporter(JSONExportConfig(format=ExportFormat.MINIMAL))
        result = exporter.export(rich_pipeline_state)
        # Minimal doesn't call _build_standard_export, so no pipeline
        assert "pipeline" not in result

    def test_detailed_format_has_pipeline_key(self, rich_pipeline_state):
        exporter = JSONExporter(JSONExportConfig(format=ExportFormat.DETAILED))
        result = exporter.export(rich_pipeline_state)
        assert "pipeline" in result

    def test_pipeline_empty_when_no_features(self, base_state):
        exporter = JSONExporter(JSONExportConfig(format=ExportFormat.STANDARD))
        result = exporter.export(base_state)
        # Pipeline should be empty dict or not present for base state
        pipeline = result.get("pipeline", {})
        # Base state has use_adaptive_extraction=True but no analyses, so extraction_mode may appear
        # The key thing: no splitting, tables, proposals, or memory
        assert "document_splitting" not in pipeline
        assert "table_detection" not in pipeline
        assert "schema_proposal" not in pipeline

    def test_pipeline_document_splitting(self, rich_pipeline_state):
        exporter = JSONExporter(JSONExportConfig(format=ExportFormat.STANDARD))
        result = exporter.export(rich_pipeline_state)
        splitting = result["pipeline"]["document_splitting"]
        assert splitting["is_multi_document"] is True
        assert splitting["segment_count"] == 2
        assert len(splitting["segments"]) == 2

    def test_pipeline_table_detection(self, rich_pipeline_state):
        exporter = JSONExporter(JSONExportConfig(format=ExportFormat.STANDARD))
        result = exporter.export(rich_pipeline_state)
        tables = result["pipeline"]["table_detection"]
        assert tables["tables_detected"] == 2
        assert len(tables["tables"]) == 2

    def test_pipeline_schema_proposal(self, rich_pipeline_state):
        exporter = JSONExporter(JSONExportConfig(format=ExportFormat.STANDARD))
        result = exporter.export(rich_pipeline_state)
        proposal = result["pipeline"]["schema_proposal"]
        assert proposal["schema_name"] == "auto_cms1500"
        assert len(proposal["fields"]) == 3

    def test_pipeline_prompt_enhancement(self, rich_pipeline_state):
        exporter = JSONExporter(JSONExportConfig(format=ExportFormat.STANDARD))
        result = exporter.export(rich_pipeline_state)
        enhancement = result["pipeline"]["prompt_enhancement"]
        assert enhancement["applied"] is True

    def test_pipeline_prompt_enhancement_not_applied(self, base_state):
        exporter = JSONExporter(JSONExportConfig(format=ExportFormat.STANDARD))
        result = exporter.export(base_state)
        pipeline = result.get("pipeline", {})
        assert "prompt_enhancement" not in pipeline

    def test_pipeline_extraction_mode(self, rich_pipeline_state):
        exporter = JSONExporter(JSONExportConfig(format=ExportFormat.STANDARD))
        result = exporter.export(rich_pipeline_state)
        mode = result["pipeline"]["extraction_mode"]
        assert mode["adaptive"] is True
        assert mode["layout_analyses"] == 2
        assert mode["component_maps"] == 1
        assert mode["adaptive_schema_generated"] is True

    def test_pipeline_memory_context(self, rich_pipeline_state):
        exporter = JSONExporter(JSONExportConfig(format=ExportFormat.STANDARD))
        result = exporter.export(rich_pipeline_state)
        memory = result["pipeline"]["memory"]
        assert memory["similar_documents"] == 2
        assert memory["correction_hints_available"] is True
        assert memory["provider_patterns_available"] is True

    def test_pipeline_memory_not_present_when_empty(self, base_state):
        exporter = JSONExporter(JSONExportConfig(format=ExportFormat.STANDARD))
        result = exporter.export(base_state)
        pipeline = result.get("pipeline", {})
        assert "memory" not in pipeline

    def test_pipeline_json_serializable(self, rich_pipeline_state, tmp_path):
        """Pipeline data must be JSON-serializable."""
        out = tmp_path / "test.json"
        exporter = JSONExporter(JSONExportConfig(format=ExportFormat.DETAILED))
        result = exporter.export(rich_pipeline_state, out)

        # Read back and verify
        with out.open() as f:
            loaded = json.load(f)
        assert "pipeline" in loaded
        assert loaded["pipeline"]["document_splitting"]["segment_count"] == 2

    def test_pipeline_with_phi_masking(self, rich_pipeline_state):
        """PHI masking should not affect pipeline section."""
        config = JSONExportConfig(format=ExportFormat.STANDARD, mask_phi=True)
        exporter = JSONExporter(config)
        result = exporter.export(rich_pipeline_state)
        # Pipeline data is not PHI — should be unaffected
        assert result["pipeline"]["document_splitting"]["is_multi_document"] is True


# ──────────────────────────────────────────────────────────────────
# Markdown: Pipeline Intelligence Section
# ──────────────────────────────────────────────────────────────────


class TestMarkdownPipelineIntelligence:
    """Tests for pipeline intelligence in Markdown exports."""

    def test_detailed_report_has_pipeline_section(self, rich_pipeline_state):
        exporter = MarkdownExporter(MarkdownExportConfig(style=MarkdownStyle.DETAILED))
        content = exporter.export(rich_pipeline_state)
        assert "Pipeline Intelligence" in content

    def test_simple_report_no_pipeline_section(self, rich_pipeline_state):
        """Simple report does not include pipeline intelligence."""
        exporter = MarkdownExporter(MarkdownExportConfig(style=MarkdownStyle.SIMPLE))
        content = exporter.export(rich_pipeline_state)
        assert "Pipeline Intelligence" not in content

    def test_summary_report_no_pipeline_section(self, rich_pipeline_state):
        exporter = MarkdownExporter(MarkdownExportConfig(style=MarkdownStyle.SUMMARY))
        content = exporter.export(rich_pipeline_state)
        assert "Pipeline Intelligence" not in content

    def test_technical_report_has_pipeline_section(self, rich_pipeline_state):
        """Technical report includes pipeline intelligence (via detailed)."""
        exporter = MarkdownExporter(MarkdownExportConfig(style=MarkdownStyle.TECHNICAL))
        content = exporter.export(rich_pipeline_state)
        assert "Pipeline Intelligence" in content

    def test_pipeline_splitting_in_markdown(self, rich_pipeline_state):
        exporter = MarkdownExporter(MarkdownExportConfig(style=MarkdownStyle.DETAILED))
        content = exporter.export(rich_pipeline_state)
        assert "Document Splitting" in content
        assert "Multi-Document" in content
        assert "Segment 1" in content
        assert "Pages 1-3" in content
        assert "CMS-1500" in content

    def test_pipeline_tables_in_markdown(self, rich_pipeline_state):
        exporter = MarkdownExporter(MarkdownExportConfig(style=MarkdownStyle.DETAILED))
        content = exporter.export(rich_pipeline_state)
        assert "Table Detection" in content
        assert "Tables Detected" in content

    def test_pipeline_prompt_enhancement_in_markdown(self, rich_pipeline_state):
        exporter = MarkdownExporter(MarkdownExportConfig(style=MarkdownStyle.DETAILED))
        content = exporter.export(rich_pipeline_state)
        assert "Prompt Enhancement" in content
        assert "Correction-Based Enhancement" in content
        assert "Applied" in content

    def test_pipeline_extraction_mode_in_markdown(self, rich_pipeline_state):
        exporter = MarkdownExporter(MarkdownExportConfig(style=MarkdownStyle.DETAILED))
        content = exporter.export(rich_pipeline_state)
        assert "Extraction Mode" in content
        assert "Adaptive (VLM-First)" in content

    def test_pipeline_memory_in_markdown(self, rich_pipeline_state):
        exporter = MarkdownExporter(MarkdownExportConfig(style=MarkdownStyle.DETAILED))
        content = exporter.export(rich_pipeline_state)
        assert "Memory Context" in content
        assert "Similar Documents" in content
        assert "Correction Hints" in content

    def test_pipeline_schema_proposal_in_markdown(self, rich_pipeline_state):
        exporter = MarkdownExporter(MarkdownExportConfig(style=MarkdownStyle.DETAILED))
        content = exporter.export(rich_pipeline_state)
        assert "Schema Proposal" in content
        assert "auto_cms1500" in content

    def test_pipeline_empty_for_base_state(self, base_state):
        """No pipeline section when all features are at defaults."""
        exporter = MarkdownExporter(MarkdownExportConfig(style=MarkdownStyle.DETAILED))
        content = exporter.export(base_state)
        # Pipeline Intelligence header should not appear if all sections are empty
        # Since adaptive is True but no analyses, the extraction_mode section may show
        # The key test: no splitting, table, proposal, or memory sections
        assert "Document Splitting" not in content
        assert "Table Detection" not in content
        assert "Schema Proposal" not in content

    def test_pipeline_written_to_file(self, rich_pipeline_state, tmp_path):
        out = tmp_path / "test.md"
        exporter = MarkdownExporter(MarkdownExportConfig(style=MarkdownStyle.DETAILED))
        content = exporter.export(rich_pipeline_state, out)
        assert out.exists()
        file_content = out.read_text(encoding="utf-8")
        assert "Pipeline Intelligence" in file_content


# ──────────────────────────────────────────────────────────────────
# Cross-Exporter Consistency
# ──────────────────────────────────────────────────────────────────


class TestCrossExporterConsistency:
    """Verify pipeline intelligence is consistent across all exporters."""

    def test_all_exporters_show_same_segment_count(self, rich_pipeline_state, tmp_path):
        # JSON
        json_result = JSONExporter(
            JSONExportConfig(format=ExportFormat.STANDARD)
        ).export(rich_pipeline_state)
        json_segments = json_result["pipeline"]["document_splitting"]["segment_count"]

        # Markdown
        md_content = MarkdownExporter(
            MarkdownExportConfig(style=MarkdownStyle.DETAILED)
        ).export(rich_pipeline_state)

        # Excel
        from openpyxl import load_workbook

        out = tmp_path / "cross.xlsx"
        ExcelExporter(
            ExcelExportConfig(
                sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence")],
                include_styling=False,
            )
        ).export(rich_pipeline_state, out)
        wb = load_workbook(out)
        ws = wb["Pipeline Intelligence"]
        rows = []
        for r in range(2, ws.max_row + 1):
            rows.append(
                (ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value)
            )
        excel_seg_count = next(r for r in rows if r[1] == "Segment Count")[2]

        # All should agree on 2 segments
        assert json_segments == 2
        assert "Segments**: 2" in md_content
        assert excel_seg_count == "2"

    def test_all_exporters_show_table_count(self, rich_pipeline_state, tmp_path):
        # JSON
        json_result = JSONExporter(
            JSONExportConfig(format=ExportFormat.STANDARD)
        ).export(rich_pipeline_state)
        json_tables = json_result["pipeline"]["table_detection"]["tables_detected"]

        # Markdown
        md_content = MarkdownExporter(
            MarkdownExportConfig(style=MarkdownStyle.DETAILED)
        ).export(rich_pipeline_state)

        # All should agree on 2 tables
        assert json_tables == 2
        assert "Tables Detected**: 2" in md_content


# ──────────────────────────────────────────────────────────────────
# Convenience Functions
# ──────────────────────────────────────────────────────────────────


class TestConvenienceFunctions:
    """Test that convenience functions still work with enhanced exporters."""

    def test_export_to_excel_convenience(self, base_state, tmp_path):
        out = tmp_path / "conv.xlsx"
        result = export_to_excel(base_state, out)
        assert result.exists()

    def test_export_to_json_convenience(self, base_state):
        result = export_to_json(base_state)
        assert "data" in result
        assert "processing_id" in result

    def test_export_to_markdown_convenience(self, base_state):
        result = export_to_markdown(base_state)
        assert "CMS-1500" in result


# ──────────────────────────────────────────────────────────────────
# Edge Cases
# ──────────────────────────────────────────────────────────────────


class TestEnhancedExportEdgeCases:
    """Test edge cases for enhanced export pipeline intelligence."""

    def test_empty_state_excel(self, tmp_path):
        """Completely empty state should not crash pipeline sheet."""
        state: dict[str, Any] = {
            "processing_id": "empty",
            "pdf_path": "test.pdf",
            "status": "pending",
            "merged_extraction": {},
            "field_metadata": {},
            "validation": {},
            "page_extractions": [],
            "page_images": [],
            "errors": [],
            "warnings": [],
        }
        config = ExcelExportConfig(
            sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline")],
            include_styling=False,
        )
        out = tmp_path / "empty.xlsx"
        ExcelExporter(config).export(state, out)
        assert out.exists()

    def test_empty_state_json(self):
        state: dict[str, Any] = {
            "processing_id": "empty",
            "pdf_path": "test.pdf",
            "status": "pending",
            "merged_extraction": {},
            "field_metadata": {},
            "page_images": [],
        }
        result = JSONExporter(JSONExportConfig(format=ExportFormat.STANDARD)).export(state)
        assert "data" in result

    def test_empty_state_markdown(self):
        state: dict[str, Any] = {
            "processing_id": "empty",
            "pdf_path": "test.pdf",
            "status": "pending",
            "document_type": "Unknown",
            "overall_confidence": 0.0,
            "merged_extraction": {},
            "field_metadata": {},
            "validation": {},
            "page_images": [],
            "page_extractions": [],
            "errors": [],
            "warnings": [],
        }
        result = MarkdownExporter(
            MarkdownExportConfig(style=MarkdownStyle.DETAILED)
        ).export(state)
        assert "Unknown Extraction Report" in result

    def test_segments_with_missing_keys(self, base_state, tmp_path):
        """Segments with missing keys should not crash."""
        base_state["is_multi_document"] = True
        base_state["document_segments"] = [
            {"start_page": 1},  # missing end_page and document_type
            {},  # completely empty
        ]
        from openpyxl import load_workbook

        config = ExcelExportConfig(
            sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline")],
            include_styling=False,
        )
        out = tmp_path / "partial.xlsx"
        ExcelExporter(config).export(base_state, out)

        wb = load_workbook(out)
        ws = wb["Pipeline"]
        # Should have rows but not crash
        assert ws.max_row > 1

    def test_tables_with_alternative_keys(self, base_state):
        """Tables using 'rows'/'columns' instead of 'row_count'/'column_count'."""
        base_state["detected_tables"] = [
            {"page": 1, "rows": 10, "columns": 3},
        ]
        result = JSONExporter(JSONExportConfig(format=ExportFormat.STANDARD)).export(base_state)
        pipeline = result.get("pipeline", {})
        assert pipeline.get("table_detection", {}).get("tables_detected") == 1

    def test_non_dict_segments_ignored(self, base_state, tmp_path):
        """Non-dict items in segments list are safely skipped."""
        base_state["is_multi_document"] = True
        base_state["document_segments"] = ["invalid", 42, None]

        config = ExcelExportConfig(
            sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline")],
            include_styling=False,
        )
        out = tmp_path / "invalid.xlsx"
        ExcelExporter(config).export(base_state, out)
        assert out.exists()

    def test_non_dict_tables_ignored(self, base_state):
        """Non-dict items in tables list are safely skipped in markdown."""
        base_state["detected_tables"] = ["invalid", 42]
        content = MarkdownExporter(
            MarkdownExportConfig(style=MarkdownStyle.DETAILED)
        ).export(base_state)
        assert "Table Detection" in content


# ──────────────────────────────────────────────────────────────────
# create_initial_state Integration
# ──────────────────────────────────────────────────────────────────


class TestInitialStateExport:
    """Test that create_initial_state produces exportable states."""

    def test_initial_state_exports_json(self):
        state = create_initial_state(pdf_path="test.pdf")
        result = JSONExporter(JSONExportConfig(format=ExportFormat.STANDARD)).export(state)
        assert result["processing_id"]
        pipeline = result.get("pipeline", {})
        # Initial state has adaptive=True but nothing else
        assert "document_splitting" not in pipeline
        assert "table_detection" not in pipeline

    def test_initial_state_exports_excel(self, tmp_path):
        state = create_initial_state(pdf_path="test.pdf")
        config = ExcelExportConfig(
            sheets=[SheetConfig(SheetType.PIPELINE, "Pipeline")],
            include_styling=False,
        )
        out = tmp_path / "initial.xlsx"
        ExcelExporter(config).export(state, out)
        assert out.exists()

    def test_initial_state_exports_markdown(self):
        state = create_initial_state(pdf_path="test.pdf")
        content = MarkdownExporter(
            MarkdownExportConfig(style=MarkdownStyle.DETAILED)
        ).export(state)
        assert "Extraction Report" in content


# ──────────────────────────────────────────────────────────────────
# Module Exports
# ──────────────────────────────────────────────────────────────────


class TestModuleExports:
    """Test module-level exports haven't regressed."""

    def test_excel_exports(self):
        from src.export import (
            SheetType,
        )

        assert SheetType.PIPELINE == "pipeline"

    def test_json_exports(self):
        from src.export import (
            ExportFormat,
        )

        assert ExportFormat.STANDARD == "standard"

    def test_markdown_exports(self):
        from src.export import (
            MarkdownStyle,
        )

        assert MarkdownStyle.DETAILED == "detailed"
