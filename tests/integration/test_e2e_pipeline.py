"""
End-to-End integration tests for the document extraction pipeline.

Tests the complete extraction workflow from PDF input
to validated, exported output.
"""

import json
import tempfile
from collections.abc import Generator
from pathlib import Path
from typing import Any

import pytest

from src.export import (
    ExportFormat,
    JSONExporter,
    export_to_excel,
    export_to_json,
)
from src.pipeline.state import (
    ConfidenceLevel,
    ExtractionStatus,
    create_initial_state,
)
from src.validation import (
    ConfidenceScorer,
    CrossFieldValidator,
    DualPassComparator,
    HallucinationPatternDetector,
)


@pytest.fixture
def temp_output_dir() -> Generator[Path, None, None]:
    """Create temporary output directory."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def sample_pdf_path(temp_output_dir: Path) -> Path:
    """Create a sample PDF file."""
    pdf_path = temp_output_dir / "sample.pdf"
    # Write minimal PDF content
    pdf_content = b"""%PDF-1.4
1 0 obj
<< /Type /Catalog /Pages 2 0 R >>
endobj
2 0 obj
<< /Type /Pages /Kids [3 0 R] /Count 1 >>
endobj
3 0 obj
<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] >>
endobj
xref
0 4
0000000000 65535 f
0000000009 00000 n
0000000058 00000 n
0000000115 00000 n
trailer
<< /Size 4 /Root 1 0 R >>
startxref
190
%%EOF"""
    pdf_path.write_bytes(pdf_content)
    return pdf_path


@pytest.fixture
def mock_extraction_result() -> dict[str, Any]:
    """Create a mock extraction result."""
    return {
        "processing_id": "e2e-test-001",
        "pdf_path": "/test/sample.pdf",
        "pdf_hash": "sha256_hash_here",
        "document_type": "CMS-1500",
        "selected_schema_name": "cms1500_v1",
        "status": ExtractionStatus.COMPLETED.value,
        "start_time": "2024-01-15T10:00:00Z",
        "end_time": "2024-01-15T10:00:45Z",
        "total_vlm_calls": 4,
        "total_processing_time_ms": 45000,
        "retry_count": 0,
        "overall_confidence": 0.89,
        "confidence_level": ConfidenceLevel.HIGH.value,
        "requires_human_review": False,
        "human_review_reason": "",
        "page_images": [b"page1_data", b"page2_data"],
        "merged_extraction": {
            "patient_name": {
                "value": "Jane Smith",
                "confidence": 0.95,
                "location": "Box 2",
            },
            "patient_dob": {
                "value": "05/12/1985",
                "confidence": 0.92,
                "location": "Box 3",
            },
            "diagnosis_code_1": {
                "value": "J06.9",
                "confidence": 0.88,
                "location": "Box 21A",
            },
            "procedure_code_1": {
                "value": "99213",
                "confidence": 0.91,
                "location": "Box 24D",
            },
            "total_charges": {
                "value": "175.00",
                "confidence": 0.93,
                "location": "Box 28",
            },
            "provider_npi": {
                "value": "1234567890",
                "confidence": 0.90,
                "location": "Box 33a",
            },
        },
        "field_metadata": {
            "patient_name": {
                "confidence": 0.95,
                "confidence_level": "high",
                "passes_agree": True,
                "validation_passed": True,
            },
            "patient_dob": {
                "confidence": 0.92,
                "confidence_level": "high",
                "passes_agree": True,
                "validation_passed": True,
            },
            "diagnosis_code_1": {
                "confidence": 0.88,
                "confidence_level": "high",
                "passes_agree": True,
                "validation_passed": True,
            },
            "procedure_code_1": {
                "confidence": 0.91,
                "confidence_level": "high",
                "passes_agree": True,
                "validation_passed": True,
            },
            "total_charges": {
                "confidence": 0.93,
                "confidence_level": "high",
                "passes_agree": True,
                "validation_passed": True,
            },
            "provider_npi": {
                "confidence": 0.90,
                "confidence_level": "high",
                "passes_agree": True,
                "validation_passed": True,
            },
        },
        "validation": {
            "is_valid": True,
            "field_validations": {
                "patient_name": {"is_valid": True},
                "diagnosis_code_1": {"is_valid": True, "validation_type": "icd10"},
                "procedure_code_1": {"is_valid": True, "validation_type": "cpt"},
                "provider_npi": {"is_valid": True, "validation_type": "npi"},
            },
            "cross_field_validations": [],
            "hallucination_flags": [],
            "warnings": [],
            "errors": [],
        },
        "page_extractions": [
            {
                "page_number": 1,
                "merged_fields": {
                    "patient_name": "Jane Smith",
                    "patient_dob": "05/12/1985",
                },
                "overall_confidence": 0.93,
                "agreement_rate": 1.0,
                "vlm_calls": 2,
                "extraction_time_ms": 22000,
                "errors": [],
                "pass1_raw": {
                    "patient_name": {"value": "Jane Smith", "confidence": 0.94},
                },
                "pass2_raw": {
                    "patient_name": {"value": "Jane Smith", "confidence": 0.95},
                },
            },
            {
                "page_number": 2,
                "merged_fields": {
                    "diagnosis_code_1": "J06.9",
                    "total_charges": "175.00",
                },
                "overall_confidence": 0.90,
                "agreement_rate": 1.0,
                "vlm_calls": 2,
                "extraction_time_ms": 23000,
                "errors": [],
                "pass1_raw": {
                    "diagnosis_code_1": {"value": "J06.9", "confidence": 0.87},
                },
                "pass2_raw": {
                    "diagnosis_code_1": {"value": "J06.9", "confidence": 0.88},
                },
            },
        ],
        "errors": [],
        "warnings": [],
    }


@pytest.mark.integration
class TestE2EPipelineFlow:
    """End-to-end tests for the complete pipeline flow."""

    def test_initial_state_creation(self, sample_pdf_path: Path) -> None:
        """Test creating initial extraction state."""
        state = create_initial_state(str(sample_pdf_path))

        assert state["pdf_path"] == str(sample_pdf_path)
        assert state["status"] == ExtractionStatus.PENDING.value
        assert "processing_id" in state
        assert state["processing_id"] != ""

    def test_json_export_from_extraction(
        self,
        mock_extraction_result: dict[str, Any],
        temp_output_dir: Path,
    ) -> None:
        """Test JSON export from extraction result."""
        output_path = temp_output_dir / "result.json"

        result = export_to_json(
            mock_extraction_result,
            output_path=output_path,
            format=ExportFormat.DETAILED,
        )

        assert output_path.exists()
        assert "data" in result
        assert result["data"]["patient_name"] == "Jane Smith"

        # Verify file content
        with open(output_path) as f:
            loaded = json.load(f)
        assert loaded["processing_id"] == "e2e-test-001"

    def test_excel_export_from_extraction(
        self,
        mock_extraction_result: dict[str, Any],
        temp_output_dir: Path,
    ) -> None:
        """Test Excel export from extraction result."""
        from openpyxl import load_workbook

        output_path = temp_output_dir / "result.xlsx"

        result_path = export_to_excel(
            mock_extraction_result,
            output_path=output_path,
        )

        assert result_path.exists()
        assert result_path.suffix == ".xlsx"

        # Verify workbook structure
        workbook = load_workbook(result_path)
        assert "Extracted Data" in workbook.sheetnames
        assert "Processing Metadata" in workbook.sheetnames

    def test_multi_format_export(
        self,
        mock_extraction_result: dict[str, Any],
        temp_output_dir: Path,
    ) -> None:
        """Test exporting to both JSON and Excel."""
        json_path = temp_output_dir / "result.json"
        excel_path = temp_output_dir / "result.xlsx"

        # Export to JSON
        json_result = export_to_json(
            mock_extraction_result,
            output_path=json_path,
            format=ExportFormat.STANDARD,
        )

        # Export to Excel
        excel_result = export_to_excel(
            mock_extraction_result,
            output_path=excel_path,
        )

        assert json_path.exists()
        assert excel_path.exists()

        # Both should have same data
        with open(json_path) as f:
            json_data = json.load(f)
        assert json_data["data"]["patient_name"] == "Jane Smith"


@pytest.mark.integration
class TestValidationIntegration:
    """Integration tests for validation components."""

    def test_dual_pass_comparison(self) -> None:
        """Test dual-pass extraction comparison."""
        from src.validation.dual_pass import ComparisonResult

        pass1_data = {
            "patient_name": "John Doe",
            "date_of_service": "01/15/2024",
        }
        pass2_data = {
            "patient_name": "John Doe",
            "date_of_service": "01/15/2024",
        }
        pass1_confidence = {"patient_name": 0.92, "date_of_service": 0.88}
        pass2_confidence = {"patient_name": 0.94, "date_of_service": 0.90}

        comparator = DualPassComparator()
        result = comparator.compare(pass1_data, pass2_data, pass1_confidence, pass2_confidence)

        assert result.overall_agreement_rate >= 0.9
        assert len(result.field_comparisons) == 2
        # Check all comparisons are matches (not MISMATCH)
        assert all(
            fc.result != ComparisonResult.MISMATCH for fc in result.field_comparisons.values()
        )

    def test_dual_pass_with_disagreement(self) -> None:
        """Test dual-pass comparison with disagreement."""
        from src.validation.dual_pass import ComparisonResult

        pass1_data = {
            "patient_name": "John Doe",
            "diagnosis_code": "J06.9",
        }
        pass2_data = {
            "patient_name": "Jane Smith",
            "diagnosis_code": "Z00.00",
        }
        pass1_confidence = {"patient_name": 0.92, "diagnosis_code": 0.75}
        pass2_confidence = {"patient_name": 0.94, "diagnosis_code": 0.72}

        comparator = DualPassComparator()
        result = comparator.compare(pass1_data, pass2_data, pass1_confidence, pass2_confidence)

        # Should detect the disagreement - completely different values
        patient_comparison = result.field_comparisons.get("patient_name")
        assert patient_comparison is not None
        # Different names should result in mismatch
        assert patient_comparison.result == ComparisonResult.MISMATCH

    def test_hallucination_pattern_detection(self) -> None:
        """Test hallucination pattern detection."""
        extraction = {
            "patient_name": {"value": "John Doe", "confidence": 0.95},
            "ssn": {"value": "123-45-6789", "confidence": 0.92},  # Suspicious pattern
            "phone": {"value": "000-000-0000", "confidence": 0.88},  # Placeholder
        }

        detector = HallucinationPatternDetector()
        result = detector.detect(extraction)

        # Should flag suspicious patterns
        assert len(result.matches) > 0

    def test_confidence_scoring(self) -> None:
        """Test confidence scoring."""
        extraction_confidences = {
            "patient_name": 0.95,
            "date_of_service": 0.88,
        }
        agreement_scores = {
            "patient_name": 1.0,
            "date_of_service": 0.95,
        }
        validation_results = {
            "patient_name": True,
            "date_of_service": True,
        }

        scorer = ConfidenceScorer()
        result = scorer.calculate(
            extraction_confidences=extraction_confidences,
            agreement_scores=agreement_scores,
            validation_results=validation_results,
        )

        assert result.overall_confidence >= 0.80

    def test_cross_field_validation(self) -> None:
        """Test cross-field validation."""
        data = {
            "service_date": "01/15/2024",
            "patient_dob": "05/12/1985",
            "total_charges": "175.00",
        }

        validator = CrossFieldValidator()
        result = validator.validate(data)

        assert result.passed is True


@pytest.mark.integration
class TestExportValidationIntegration:
    """Integration tests for export with validation."""

    def test_export_includes_validation(
        self,
        mock_extraction_result: dict[str, Any],
        temp_output_dir: Path,
    ) -> None:
        """Test that export includes validation results."""
        output_path = temp_output_dir / "result.json"

        result = export_to_json(
            mock_extraction_result,
            output_path=output_path,
            format=ExportFormat.DETAILED,
        )

        assert "validation" in result
        assert result["validation"]["is_valid"] is True

    def test_export_phi_masking(
        self,
        mock_extraction_result: dict[str, Any],
        temp_output_dir: Path,
    ) -> None:
        """Test PHI masking in exports."""
        from src.export.json_exporter import JSONExportConfig

        config = JSONExportConfig(
            format=ExportFormat.STANDARD,
            mask_phi=True,
            phi_fields={"patient_name", "patient_dob"},
        )

        exporter = JSONExporter(config)
        result = exporter.export(mock_extraction_result)

        # PHI fields should be masked
        assert "***MASKED***" in str(result["data"]["patient_name"])
        assert "***MASKED***" in str(result["data"]["patient_dob"])

        # Non-PHI fields should not be masked
        assert result["data"]["diagnosis_code_1"] == "J06.9"

    def test_export_field_exclusion(
        self,
        mock_extraction_result: dict[str, Any],
    ) -> None:
        """Test field exclusion in exports."""
        from src.export.json_exporter import JSONExportConfig

        config = JSONExportConfig(
            format=ExportFormat.MINIMAL,
            exclude_fields={"provider_npi", "total_charges"},
        )

        exporter = JSONExporter(config)
        result = exporter.export(mock_extraction_result)

        assert "provider_npi" not in result["data"]
        assert "total_charges" not in result["data"]
        assert "patient_name" in result["data"]


@pytest.mark.integration
class TestPipelineStateTransitions:
    """Test pipeline state transitions."""

    def test_state_progression(self, sample_pdf_path: Path) -> None:
        """Test state progresses through expected stages."""
        state = create_initial_state(str(sample_pdf_path))

        # Initial state
        assert state["status"] == ExtractionStatus.PENDING.value

        # Update through stages
        from src.pipeline.state import update_state

        state = update_state(state, {"status": ExtractionStatus.PREPROCESSING.value})
        assert state["status"] == ExtractionStatus.PREPROCESSING.value

        state = update_state(state, {"status": ExtractionStatus.EXTRACTING.value})
        assert state["status"] == ExtractionStatus.EXTRACTING.value

        state = update_state(state, {"status": ExtractionStatus.VALIDATING.value})
        assert state["status"] == ExtractionStatus.VALIDATING.value

        state = update_state(state, {"status": ExtractionStatus.COMPLETED.value})
        assert state["status"] == ExtractionStatus.COMPLETED.value

    def test_failed_state(self, sample_pdf_path: Path) -> None:
        """Test failure state handling."""
        from src.pipeline.state import update_state

        state = create_initial_state(str(sample_pdf_path))
        state = update_state(
            state,
            {
                "status": ExtractionStatus.FAILED.value,
                "errors": ["Processing error occurred"],
            },
        )

        assert state["status"] == ExtractionStatus.FAILED.value
        assert len(state["errors"]) == 1

    def test_human_review_state(self, sample_pdf_path: Path) -> None:
        """Test human review state."""
        from src.pipeline.state import update_state

        state = create_initial_state(str(sample_pdf_path))
        state = update_state(
            state,
            {
                "status": ExtractionStatus.HUMAN_REVIEW.value,
                "requires_human_review": True,
                "human_review_reason": "Low confidence on critical fields",
            },
        )

        assert state["status"] == ExtractionStatus.HUMAN_REVIEW.value
        assert state["requires_human_review"] is True
        assert "Low confidence" in state["human_review_reason"]


@pytest.mark.integration
class TestAPIIntegration:
    """Integration tests for API with pipeline."""

    def test_api_process_response_format(
        self,
        mock_extraction_result: dict[str, Any],
    ) -> None:
        """Test API response format matches expected schema."""
        from src.api.routes.documents import _build_process_response

        response = _build_process_response(mock_extraction_result, "/output/result.json")

        assert response.processing_id == "e2e-test-001"
        assert response.status.value == "completed"
        assert response.overall_confidence == 0.89
        assert "patient_name" in response.data
        assert response.output_path == "/output/result.json"

    def test_api_validation_in_response(
        self,
        mock_extraction_result: dict[str, Any],
    ) -> None:
        """Test validation results in API response."""
        from src.api.routes.documents import _build_process_response

        response = _build_process_response(mock_extraction_result)

        assert response.validation is not None
        assert response.validation.is_valid is True


@pytest.mark.integration
class TestQueueIntegration:
    """Integration tests for queue with pipeline."""

    def test_task_result_format(self) -> None:
        """Test task result format."""
        from src.queue.tasks import TaskResult, TaskStatus

        result = TaskResult(
            task_id="task-123",
            processing_id="proc-001",
            status=TaskStatus.COMPLETED,
            field_count=6,
            overall_confidence=0.89,
        )

        result_dict = result.to_dict()

        assert result_dict["task_id"] == "task-123"
        assert result_dict["status"] == "completed"
        assert result_dict["field_count"] == 6

    def test_task_result_from_extraction(
        self,
        mock_extraction_result: dict[str, Any],
    ) -> None:
        """Test creating task result from extraction."""
        from src.queue.tasks import TaskResult, TaskStatus

        result = TaskResult(
            task_id="task-456",
            processing_id=mock_extraction_result["processing_id"],
            status=TaskStatus.COMPLETED,
            field_count=len(mock_extraction_result["merged_extraction"]),
            overall_confidence=mock_extraction_result["overall_confidence"],
        )

        assert result.processing_id == "e2e-test-001"
        assert result.field_count == 6
        assert result.overall_confidence == 0.89
